"""
Reports Module — PDF & Excel Generators
Simple version (no categories, no structured items)
"""

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


# ── PDF: Patient Visit Report ──────────────────────────────────────────

def generate_patient_visit_report(patient, encounters):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"Patient Visit Report", styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))

    # Patient Info
    elements.append(Paragraph(f"<b>Patient:</b> {patient.user.get_full_name()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Patient Number:</b> {patient.patient_number}", styles['Normal']))
    elements.append(Paragraph(f"<b>DOB:</b> {patient.date_of_birth}", styles['Normal']))
    elements.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))

    # Encounters Table
    data = [['Date', 'Doctor', 'Chief Complaint', 'Status']]
    for enc in encounters:
        data.append([
            enc.started_at.strftime('%Y-%m-%d %H:%M'),
            enc.doctor.user.get_full_name() if enc.doctor else 'N/A',
            enc.chief_complaint[:50] + '...' if len(enc.chief_complaint) > 50 else enc.chief_complaint,
            enc.get_status_display()
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ── PDF: Lab Report (simple) ──────────────────────────────────────────

def generate_lab_report(patient, lab_orders):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"Lab Results Report", styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))

    # Patient Info
    elements.append(Paragraph(f"<b>Patient:</b> {patient.user.get_full_name()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Patient Number:</b> {patient.patient_number}", styles['Normal']))
    elements.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))

    # Lab Orders Table
    data = [['Order Date', 'Test', 'Status', 'Result']]
    for order in lab_orders:
        data.append([
            order.ordered_at.strftime('%Y-%m-%d'),
            order.test.name,
            order.get_status_display(),
            order.result_text[:50] + '...' if order.result_text and len(order.result_text) > 50 else order.result_text or 'Pending'
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ── PDF: Radiology Report (simple) ─────────────────────────────────────

def generate_radiology_report(patient, radiology_orders):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"Radiology Report", styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))

    # Patient Info
    elements.append(Paragraph(f"<b>Patient:</b> {patient.user.get_full_name()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Patient Number:</b> {patient.patient_number}", styles['Normal']))
    elements.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))

    # Radiology Orders Table
    data = [['Order Date', 'Test', 'Status', 'Result']]
    for order in radiology_orders:
        data.append([
            order.ordered_at.strftime('%Y-%m-%d'),
            order.test.name,
            order.get_status_display(),
            order.result_text[:50] + '...' if order.result_text and len(order.result_text) > 50 else order.result_text or 'Pending'
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ── PDF: Prescription Report ────────────────────────────────────────────
def generate_prescription_report(patient, prescriptions):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"Prescription History Report", styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))

    # Patient Info
    elements.append(Paragraph(f"<b>Patient:</b> {patient.user.get_full_name()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Patient Number:</b> {patient.patient_number}", styles['Normal']))
    elements.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))

    # Prescriptions Table
    data = [['Date', 'Medicine', 'Dosage', 'Frequency', 'Status']]
    for prescription in prescriptions:
        for item in prescription.prescriptionitem_set.all():
            data.append([
                prescription.prescribed_at.strftime('%Y-%m-%d'),
                item.medicine.name,
                item.dosage,
                item.frequency,
                prescription.get_status_display()
            ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer



# ── Excel: Patient Visit Report ─────────────────────────────────────────

def generate_patient_visit_excel(patient, encounters):
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Visits"

    ws['A1'] = "Patient Visit Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Patient: {patient.user.get_full_name()}"
    ws['A3'] = f"Patient Number: {patient.patient_number}"
    ws['A4'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    headers = ['Date', 'Doctor', 'Chief Complaint', 'Status']
    for i, h in enumerate(headers):
        cell = ws.cell(row=7, column=i+1, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row, enc in enumerate(encounters, start=8):
        ws.cell(row=row, column=1, value=enc.started_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=enc.doctor.user.get_full_name() if enc.doctor else 'N/A')
        ws.cell(row=row, column=3, value=enc.chief_complaint)
        ws.cell(row=row, column=4, value=enc.get_status_display())

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer